#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Local-only document text extraction and SQLite full-text search."""
from __future__ import annotations

import json
import os
import re
import sqlite3
import threading
import time
import zipfile
from pathlib import Path
from typing import Any, Dict, Iterable, List
from xml.etree import ElementTree

import fitz
from openpyxl import load_workbook

from api.utils.error_handler import api_error, api_success
from pyapp.config.config import Config

INDEX_EXTENSIONS = {'.pdf', '.docx', '.xlsx', '.xlsm', '.txt', '.md', '.markdown', '.csv', '.json', '.log'}
_MAX_TEXT_BYTES = 8 * 1024 * 1024
_MAX_CONTENT_CHARS = 2_000_000


class DocumentIndexMixin:
    """Persistent document index stored under the app data directory."""

    _document_index_boot_lock = threading.Lock()

    def _document_index_ensure(self) -> None:
        if getattr(self, '_document_index_ready', False):
            return
        with self._document_index_boot_lock:
            if getattr(self, '_document_index_ready', False):
                return
            index_dir = Path(Config.appDataDir) / 'document-index'
            index_dir.mkdir(parents=True, exist_ok=True)
            self._document_index_path = index_dir / 'documents.sqlite3'
            self._document_index_lock = threading.RLock()
            with self._document_index_connect() as connection:
                connection.execute('PRAGMA journal_mode=WAL')
                connection.execute('PRAGMA synchronous=NORMAL')
                connection.execute(
                    '''CREATE TABLE IF NOT EXISTS documents (
                        path TEXT PRIMARY KEY,
                        title TEXT NOT NULL,
                        extension TEXT NOT NULL,
                        content TEXT NOT NULL,
                        mtime_ns INTEGER NOT NULL,
                        size INTEGER NOT NULL,
                        indexed_at REAL NOT NULL
                    )'''
                )
                connection.execute('CREATE TABLE IF NOT EXISTS index_meta (key TEXT PRIMARY KEY, value TEXT NOT NULL)')
                self._document_index_create_fts(connection)
                connection.commit()
            self._document_index_ready = True

    def _document_index_connect(self):
        return sqlite3.connect(str(self._document_index_path), timeout=30)

    @staticmethod
    def _document_index_create_fts(connection) -> None:
        exists = connection.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name='documents_fts'").fetchone()
        if exists:
            return
        tokenizer = 'trigram'
        try:
            connection.execute("CREATE VIRTUAL TABLE documents_fts USING fts5(path UNINDEXED, title, content, tokenize='trigram')")
        except sqlite3.OperationalError:
            tokenizer = 'unicode61'
            connection.execute("CREATE VIRTUAL TABLE documents_fts USING fts5(path UNINDEXED, title, content, tokenize='unicode61')")
        connection.execute('INSERT OR REPLACE INTO index_meta(key, value) VALUES (?, ?)', ('tokenizer', tokenizer))

    @staticmethod
    def _document_index_safe_path(raw: Any, expect_directory=False) -> Path:
        path = Path(str(raw or '')).expanduser().resolve()
        if expect_directory and not path.is_dir():
            raise ValueError(f'目录不存在：{path}')
        if not expect_directory and not path.is_file():
            raise ValueError(f'文件不存在：{path}')
        return path

    @staticmethod
    def _document_index_extract_docx(path: Path) -> str:
        with zipfile.ZipFile(path) as archive:
            raw = archive.read('word/document.xml')
        root = ElementTree.fromstring(raw)
        namespace = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
        paragraphs = []
        for paragraph in root.iter(f'{namespace}p'):
            text = ''.join(node.text or '' for node in paragraph.iter(f'{namespace}t'))
            if text.strip():
                paragraphs.append(text)
        return '\n'.join(paragraphs)

    @staticmethod
    def _document_index_extract_xlsx(path: Path) -> str:
        workbook = load_workbook(path, read_only=True, data_only=True)
        chunks = []
        try:
            for sheet in workbook.worksheets:
                chunks.append(f'--- {sheet.title} ---')
                for row in sheet.iter_rows(values_only=True):
                    values = [str(value) for value in row if value is not None and str(value).strip()]
                    if values:
                        chunks.append('\t'.join(values))
                    if sum(len(item) for item in chunks) >= _MAX_CONTENT_CHARS:
                        break
        finally:
            workbook.close()
        return '\n'.join(chunks)

    @classmethod
    def _document_index_extract(cls, path: Path) -> str:
        extension = path.suffix.lower()
        if extension == '.pdf':
            with fitz.open(path) as doc:
                return '\n\n'.join(page.get_text('text') or '' for page in doc)
        if extension == '.docx':
            return cls._document_index_extract_docx(path)
        if extension in {'.xlsx', '.xlsm'}:
            return cls._document_index_extract_xlsx(path)
        if path.stat().st_size > _MAX_TEXT_BYTES:
            raise ValueError('文本文件超过 8 MB 限制')
        try:
            return path.read_text(encoding='utf-8')
        except UnicodeDecodeError:
            return path.read_text(encoding='utf-8', errors='replace')

    @staticmethod
    def _document_index_walk(directories: Iterable[Path], recursive: bool, extensions: set) -> Iterable[Path]:
        for directory in directories:
            pattern = '**/*' if recursive else '*'
            for path in directory.glob(pattern):
                if not path.is_file() or path.suffix.lower() not in extensions:
                    continue
                if any(part.startswith('.') or part in {'node_modules', '__pycache__'} for part in path.relative_to(directory).parts[:-1]):
                    continue
                yield path.resolve()

    @staticmethod
    def _document_index_source_state(path: str, mtime_ns: int, size: int) -> Dict[str, bool]:
        try:
            stat = Path(path).stat()
        except OSError:
            return {'stale': True, 'missing': True}
        return {
            'stale': stat.st_mtime_ns != int(mtime_ns) or stat.st_size != int(size),
            'missing': False,
        }

    def document_index_status(self):
        try:
            self._document_index_ensure()
            with self._document_index_lock, self._document_index_connect() as connection:
                count, total_size, updated_at = connection.execute(
                    'SELECT COUNT(*), COALESCE(SUM(size), 0), COALESCE(MAX(indexed_at), 0) FROM documents'
                ).fetchone()
                tokenizer_row = connection.execute("SELECT value FROM index_meta WHERE key='tokenizer'").fetchone()
                sources = connection.execute('SELECT path, mtime_ns, size FROM documents').fetchall()
            stale_samples = []
            stale_documents = 0
            missing_documents = 0
            for path, mtime_ns, size in sources:
                source_state = self._document_index_source_state(path, mtime_ns, size)
                if source_state['stale']:
                    stale_documents += 1
                    missing_documents += int(source_state['missing'])
                    if len(stale_samples) < 20:
                        stale_samples.append({
                            'path': path,
                            'reason': 'missing' if source_state['missing'] else 'changed',
                        })
            return api_success(
                documents=count,
                freshDocuments=max(0, count - stale_documents),
                staleDocuments=stale_documents,
                missingDocuments=missing_documents,
                staleSamples=stale_samples,
                sourceBytes=total_size,
                databaseBytes=self._document_index_path.stat().st_size if self._document_index_path.exists() else 0,
                updatedAt=updated_at,
                tokenizer=tokenizer_row[0] if tokenizer_row else 'unicode61',
                path=str(self._document_index_path),
                extensions=sorted(INDEX_EXTENSIONS),
            )
        except Exception as exc:
            return api_error(f'读取索引状态失败：{exc}')

    def document_index_build(self, options: Dict | None = None):
        try:
            self._document_index_ensure()
            options = options or {}
            raw_directories = options.get('directories') or []
            raw_files = options.get('files') or []
            if isinstance(raw_directories, str):
                raw_directories = [raw_directories]
            if isinstance(raw_files, str):
                raw_files = [raw_files]
            directories = [self._document_index_safe_path(item, True) for item in raw_directories if str(item).strip()]
            files = [self._document_index_safe_path(item) for item in raw_files if str(item).strip()]
            if not directories and not files:
                raise ValueError('请选择要建立索引的目录或文件')
            requested_extensions = options.get('extensions') or []
            if isinstance(requested_extensions, str):
                requested_extensions = [item.strip() for item in requested_extensions.split(',') if item.strip()]
            extensions = {f'.{str(item).lower().lstrip(".")}' for item in requested_extensions} if requested_extensions else set(INDEX_EXTENSIONS)
            extensions &= INDEX_EXTENSIONS
            candidates = {path for path in files if path.suffix.lower() in extensions}
            candidates.update(self._document_index_walk(directories, bool(options.get('recursive', True)), extensions))

            indexed = 0
            skipped = 0
            failed = []
            with self._document_index_lock, self._document_index_connect() as connection:
                if bool(options.get('rebuild', False)):
                    connection.execute('DELETE FROM documents')
                    connection.execute('DELETE FROM documents_fts')
                existing = {
                    row[0]: (row[1], row[2])
                    for row in connection.execute('SELECT path, mtime_ns, size FROM documents')
                }
                for path in sorted(candidates):
                    try:
                        stat = path.stat()
                        key = str(path)
                        if existing.get(key) == (stat.st_mtime_ns, stat.st_size):
                            skipped += 1
                            continue
                        content = self._document_index_extract(path)[:_MAX_CONTENT_CHARS]
                        title = path.stem
                        connection.execute(
                            '''INSERT INTO documents(path, title, extension, content, mtime_ns, size, indexed_at)
                               VALUES (?, ?, ?, ?, ?, ?, ?)
                               ON CONFLICT(path) DO UPDATE SET title=excluded.title, extension=excluded.extension,
                               content=excluded.content, mtime_ns=excluded.mtime_ns, size=excluded.size,
                               indexed_at=excluded.indexed_at''',
                            (key, title, path.suffix.lower(), content, stat.st_mtime_ns, stat.st_size, time.time()),
                        )
                        connection.execute('DELETE FROM documents_fts WHERE path = ?', (key,))
                        connection.execute('INSERT INTO documents_fts(path, title, content) VALUES (?, ?, ?)', (key, title, content))
                        indexed += 1
                    except Exception as exc:
                        failed.append({'path': str(path), 'error': str(exc)})
                if bool(options.get('prune', False)):
                    live = {str(path) for path in candidates}
                    for key in existing:
                        if key not in live and any(key == str(directory) or key.startswith(f'{directory}{os.sep}') for directory in directories):
                            connection.execute('DELETE FROM documents WHERE path = ?', (key,))
                            connection.execute('DELETE FROM documents_fts WHERE path = ?', (key,))
                connection.commit()
            return api_success(
                f'索引更新完成：新增或更新 {indexed} 个，跳过 {skipped} 个，失败 {len(failed)} 个',
                indexed=indexed,
                skipped=skipped,
                failed=failed[:50],
                scanned=len(candidates),
            )
        except Exception as exc:
            return api_error(f'建立索引失败：{exc}')

    @staticmethod
    def _document_index_match_query(query: str) -> str:
        tokens = [token for token in re.split(r'\s+', query.strip()) if token]
        return ' AND '.join(f'"{token.replace(chr(34), chr(34) * 2)}"' for token in tokens)

    def document_index_search(self, options: Dict | str | None = None):
        try:
            self._document_index_ensure()
            if isinstance(options, str):
                options = {'query': options}
            options = options or {}
            query = str(options.get('query') or '').strip()
            if not query:
                return api_error('请输入搜索关键词')
            limit = max(1, min(int(options.get('limit') or 50), 200))
            extension = str(options.get('extension') or '').lower().strip()
            directory = str(options.get('directory') or '').strip()
            with self._document_index_lock, self._document_index_connect() as connection:
                tokenizer_row = connection.execute("SELECT value FROM index_meta WHERE key='tokenizer'").fetchone()
                tokenizer = tokenizer_row[0] if tokenizer_row else 'unicode61'
                use_like = tokenizer == 'trigram' and any(len(token) < 3 for token in re.split(r'\s+', query) if token)
                if not use_like:
                    try:
                        rows = connection.execute(
                            '''SELECT f.path, f.title, d.extension,
                                      snippet(documents_fts, 2, '<mark>', '</mark>', '…', 28) AS excerpt,
                                      bm25(documents_fts) AS score, d.size, d.indexed_at, d.mtime_ns
                               FROM documents_fts AS f
                               JOIN documents AS d ON d.path = f.path
                               WHERE documents_fts MATCH ?
                                 AND (? = '' OR d.extension = ?)
                                 AND (? = '' OR d.path LIKE ?)
                               ORDER BY score LIMIT ?''',
                            (self._document_index_match_query(query), extension, extension, directory, f'{directory}%', limit),
                        ).fetchall()
                    except sqlite3.OperationalError:
                        use_like = True
                if use_like:
                    rows = connection.execute(
                        '''SELECT path, title, extension,
                                  substr(content, max(1, instr(lower(content), lower(?)) - 80), 240),
                                  0, size, indexed_at, mtime_ns
                           FROM documents
                           WHERE (lower(title) LIKE lower(?) OR lower(content) LIKE lower(?))
                             AND (? = '' OR extension = ?)
                             AND (? = '' OR path LIKE ?)
                           ORDER BY indexed_at DESC LIMIT ?''',
                        (query, f'%{query}%', f'%{query}%', extension, extension, directory, f'{directory}%', limit),
                    ).fetchall()
            results = []
            for row in rows:
                source_state = self._document_index_source_state(row[0], row[7], row[5])
                results.append({
                    'path': row[0],
                    'title': row[1],
                    'extension': row[2],
                    'excerpt': row[3] or '',
                    'score': row[4],
                    'size': row[5],
                    'indexedAt': row[6],
                    **source_state,
                })
            return api_success(f'找到 {len(results)} 个结果', results=results, query=query)
        except Exception as exc:
            return api_error(f'搜索失败：{exc}')

    def document_index_remove(self, options: Dict | str | None = None):
        try:
            self._document_index_ensure()
            path = str(options.get('path') if isinstance(options, dict) else options or '')
            with self._document_index_lock, self._document_index_connect() as connection:
                connection.execute('DELETE FROM documents WHERE path = ?', (path,))
                connection.execute('DELETE FROM documents_fts WHERE path = ?', (path,))
                connection.commit()
            return api_success('已从索引移除', path=path)
        except Exception as exc:
            return api_error(f'移除索引失败：{exc}')

    def document_index_clear(self):
        try:
            self._document_index_ensure()
            with self._document_index_lock, self._document_index_connect() as connection:
                connection.execute('DELETE FROM documents')
                connection.execute('DELETE FROM documents_fts')
                connection.commit()
            return api_success('本地文档索引已清空')
        except Exception as exc:
            return api_error(f'清空索引失败：{exc}')
