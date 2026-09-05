#!/usr/bin/env python3
# -*- coding: utf-8 -*-
'''
Author: Codex
Date: 2025-11-11
Description: Excel 工具相关 API
'''

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import coordinate_from_string, get_column_letter

from api.core.context import checkpoint, report_progress
from api.core.outputs import atomic_output, write_output
from api.utils.validators import ensure_output_directory


class Excel():
    '''Excel 相关功能'''

    _supported_suffix = ('.xlsx', '.xlsm', '.xltx', '.xltm')
    _unsafe_filename_chars = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

    # -------- helpers --------

    def _validate_payload(self, options: Optional[Dict]) -> Dict:
        if options is None:
            return {}
        if not isinstance(options, dict):
            raise ValueError('参数格式错误')
        return options

    def _ensure_excel_file(self, file_path: str) -> Path:
        if not file_path:
            raise ValueError('请选择 Excel 文件')
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f'文件不存在：{path}')
        if path.suffix.lower() not in self._supported_suffix:
            raise ValueError('仅支持 .xlsx / .xlsm / .xltx / .xltm 文件')
        return path

    def _load_sheet(self, file_path: Path, sheet_name: Optional[str], options=None, limit=None, offset=0) -> Tuple[List[str], List[List[Any]], str]:
        options = options or {}
        header_row = max(1, int(options.get('headerRow') or 1))
        policy = options.get('formulaPolicy') or 'preserve'
        if policy not in {'preserve', 'values'}:
            raise ValueError('公式策略必须是 preserve 或 values')
        wb = load_workbook(file_path, read_only=True, data_only=False)
        cached = load_workbook(file_path, read_only=True, data_only=True) if policy == 'values' else None
        try:
            if sheet_name and sheet_name not in wb.sheetnames:
                raise ValueError(f'工作表不存在：{sheet_name}')
            ws = wb[sheet_name] if sheet_name else wb.active
            header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), ())
            header = self._normalize_schema('', '|', [self._normalize_cell(cell) for cell in header_cells])
            data_rows = []
            start = header_row + 1 + max(0, offset)
            maximum = start + limit - 1 if limit is not None else None
            cached_rows = iter(cached[ws.title].iter_rows(min_row=start, max_row=maximum)) if cached else None
            for row_number, row in enumerate(ws.iter_rows(min_row=start, max_row=maximum), start):
                cached_row = next(cached_rows, ()) if cached_rows else ()
                if row_number % 500 == 0:
                    report_progress(row_number - start, max(1, (ws.max_row or row_number) - start), '正在读取表格')
                if all(cell.value is None or cell.value == '' for cell in row):
                    continue
                cells = []
                for index, cell in enumerate(row):
                    value = cell.value
                    kind = cell.data_type
                    if policy == 'values' and kind == 'f':
                        value = cached_row[index].value if index < len(cached_row) else None
                        if value is None:
                            raise ValueError(f'{cell.coordinate} 的公式没有缓存值，请先用 Excel 或 LibreOffice 计算并保存')
                        kind = cached_row[index].data_type
                    cells.append({'value': value, 'format': cell.number_format, 'type': kind, 'coordinate': getattr(cell, 'coordinate', '')})
                data_rows.append(cells)
            return header, data_rows, ws.title
        finally:
            wb.close()
            if cached:
                cached.close()

    def _normalize_cell(self, value: Any) -> str:
        if value is None:
            return ''
        if isinstance(value, (date, datetime)):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        return str(value).strip()

    def _normalize_schema(self, schema_text: str, delimiter: str, header: Sequence[str]) -> List[str]:
        if schema_text:
            chunks = [part.strip() for part in schema_text.split(delimiter or '|')]
            normalized = [chunk for chunk in chunks if chunk]
            if normalized:
                header = normalized
        # Fallback to workbook header
        fallback = [col or f'列{i + 1}' for i, col in enumerate(header)]
        if not fallback:
            fallback = ['列1']
        seen = set()
        unique = []
        for name in fallback:
            base, suffix = name, 2
            while name in seen or name in {'_source', '_cells'}:
                name = f'{base}_{suffix}'
                suffix += 1
            seen.add(name)
            unique.append(name)
        return unique

    def _rows_to_dicts(self, schema: Sequence[str], rows: Iterable[Sequence[Any]], source: str) -> List[Dict[str, Any]]:
        mapped: List[Dict[str, Any]] = []
        for raw in rows:
            row_dict: Dict[str, Any] = {}
            metadata = {}
            for idx, key in enumerate(schema):
                cell = raw[idx] if idx < len(raw) else None
                if isinstance(cell, dict) and 'value' in cell:
                    row_dict[key] = cell['value']
                    metadata[key] = cell
                else:
                    row_dict[key] = cell
            row_dict['_source'] = source
            row_dict['_cells'] = metadata
            mapped.append(row_dict)
        return mapped

    def _ensure_output_dir(self, source: Path, preferred: str, suffix: str) -> Path:
        # 复用通用输出目录创建逻辑，保持与其它模块一致
        return ensure_output_directory(source, preferred, suffix)

    def _write_rows(self, schema: Sequence[str], rows: Iterable[Dict[str, Any]], dest: Path):
        wb = Workbook(write_only=True)
        ws = wb.create_sheet('Sheet1')
        try:
            ws.append(list(schema))
            for row_index, row in enumerate(rows, 2):
                checkpoint()
                cells = []
                for column_index, column in enumerate(schema, 1):
                    metadata = row.get('_cells', {}).get(column, {})
                    cell = WriteOnlyCell(ws, value=row.get(column, ''))
                    if metadata:
                        cell.number_format = metadata.get('format') or 'General'
                        if metadata.get('type') == 'f' and isinstance(cell.value, str):
                            origin_column, origin_row = coordinate_from_string(metadata['coordinate'])
                            if origin_column != get_column_letter(column_index):
                                raise ValueError('包含公式的列改变了位置，请使用缓存值策略或先在 Excel 中转换为数值')
                            for token in Tokenizer(cell.value).items:
                                if token.type == 'OPERAND' and token.subtype == 'RANGE':
                                    for reference in token.value.split(':'):
                                        match = re.fullmatch(r'\$?[A-Za-z]+\$?(\d+)', reference)
                                        if not match or int(match.group(1)) != origin_row or '$' in reference:
                                            raise ValueError('保留公式仅支持本行相对引用；跨行、跨表、命名区域或绝对引用请先在 Excel 中计算并选择缓存值策略')
                            cell.value = Translator(cell.value, origin=metadata['coordinate']).translate_formula(f'{get_column_letter(column_index)}{row_index}')
                        elif metadata.get('type') == 's':
                            cell.data_type = 's'
                    cells.append(cell)
                ws.append(cells)
            with atomic_output(dest) as (temporary, final):
                wb.save(temporary)
            return final
        finally:
            if not ws.closed:
                ws.close()
            writer = getattr(ws, '_writer', None)
            if writer and Path(writer.out).exists():
                writer.cleanup()
            wb.close()

    @staticmethod
    def _sort_key(value):
        if value is None or value == '':
            return (3, '')
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return (0, value)
        if isinstance(value, (date, datetime)):
            return (1, value.isoformat())
        return (2, str(value).casefold())

    @staticmethod
    def _public_rows(rows):
        return [{key: value.isoformat() if isinstance(value, (date, datetime)) else value
                 for key, value in row.items() if not key.startswith('_')} for row in rows]

    def _safe_filename_part(self, value: Any, fallback: str) -> str:
        safe_value = self._unsafe_filename_chars.sub('_', self._normalize_cell(value))
        safe_value = safe_value.strip().rstrip('. ')
        if safe_value in {'', '.', '..'}:
            return fallback
        return safe_value[:120].rstrip('. ') or fallback

    def _write_groups(self, schema: Sequence[str], grouped: Dict[str, List[Dict[str, Any]]], base_dir: Path, stem: str) -> List[str]:
        base_dir.mkdir(parents=True, exist_ok=True)
        base_dir = base_dir.resolve()
        safe_stem = self._safe_filename_part(stem, '导出')
        used_filenames = set()
        exports: List[str] = []
        for key, rows in grouped.items():
            safe_key = self._safe_filename_part(key, '未分组')
            duplicate_index = 1
            while True:
                duplicate_suffix = '' if duplicate_index == 1 else f'_{duplicate_index}'
                unique_key = f'{safe_key[:max(1, 120 - len(duplicate_suffix))]}{duplicate_suffix}'
                filename = f'{safe_stem}_{unique_key}.xlsx'
                if filename.casefold() not in used_filenames:
                    break
                duplicate_index += 1

            dest = (base_dir / filename).resolve()
            try:
                dest.relative_to(base_dir)
            except ValueError as exc:
                raise ValueError('分组导出路径无效') from exc
            dest = self._write_rows(schema, rows, dest)
            used_filenames.add(filename.casefold())
            exports.append(str(dest))
        return exports

    def _write_chart_json(self, chart: Dict[str, Any], output_dir: Path, stem: str) -> str:
        json_path = output_dir / f'{stem}_chart.json'
        json_path = write_output(json_path, lambda target: target.write_text(json.dumps(chart, ensure_ascii=False, indent=2), encoding='utf-8'))
        return str(json_path)

    def _build_chart(self, groups: Dict[str, List[Dict[str, Any]]]) -> Dict[str, Any]:
        labels: List[str] = []
        values: List[int] = []
        for key, rows in groups.items():
            labels.append(key or '未分组')
            values.append(len(rows))
        return {
            'labels': labels,
            'datasets': [
                {
                    'label': '分组数量',
                    'data': values
                }
            ],
            'generatedAt': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

    def _coerce_number(self, value: Any) -> Optional[float]:
        if value is None:
            return None
        try:
            return float(str(value).replace(',', '').strip())
        except Exception:
            return None

    def excel_preview(self, options: Dict = None):
        '''读取 Excel 表头与样例数据'''
        try:
            opts = self._validate_payload(options)
            file_path = opts.get('filePath', '')
            sheet_name = opts.get('sheetName')
            delimiter = opts.get('delimiter') or '|'

            source = self._ensure_excel_file(file_path)
            sheets = self._list_sheets(source)
            limit = max(1, min(int(opts.get('limit') or 30), 200))
            offset = max(0, int(opts.get('offset') or 0))
            header, data_rows, active_sheet = self._load_sheet(source, sheet_name, opts, limit=limit, offset=offset)
            schema = self._normalize_schema(opts.get('schemaText', ''), delimiter, header)
            rows = self._rows_to_dicts(schema, data_rows, active_sheet)
            wb = load_workbook(source, read_only=True, data_only=True)
            try:
                ws = wb[active_sheet]
                row_count = max(0, (ws.max_row or 0) - max(1, int(opts.get('headerRow') or 1)))
            finally:
                wb.close()

            return {
                'code': 0,
                'msg': '解析完成',
                'schema': schema,
                'schemaText': delimiter.join(schema),
                'delimiter': delimiter,
                'rowCount': row_count,
                'rowCountEstimated': True,
                'offset': offset,
                'limit': limit,
                'hasMore': offset + limit < row_count,
                'formulaPolicy': opts.get('formulaPolicy') or 'preserve',
                'sheet': active_sheet,
                'sheets': sheets,
                'sample': self._public_rows(rows),
                'cellTypes': [{key: value.get('type') for key, value in row.get('_cells', {}).items()} for row in rows],
                'formulaRule': 'preserve 保留本行相对引用并随排序平移；values 使用 Excel 上次保存的缓存值，不重新计算公式。输出为新 xlsx 数据表，不包含宏。'
            }
        except Exception as exc:
            return {'code': -1, 'msg': f'解析失败：{exc}'}

    def excel_process(self, options: Dict = None):
        '''执行分组、排序与导出'''
        try:
            opts = self._validate_payload(options)
            source = self._ensure_excel_file(opts.get('filePath', ''))
            sheet_name = opts.get('sheetName')
            delimiter = opts.get('delimiter') or '|'
            schema_text = opts.get('schemaText', '')
            group_by = opts.get('groupBy') or ''
            sort_by = opts.get('sortBy') or ''
            sort_order = (opts.get('sortOrder') or 'asc').lower()
            export_groups = bool(opts.get('exportGroups', True))
            export_json = bool(opts.get('exportJson', True))
            export_combined = bool(opts.get('exportCombined', False))
            output_dir = self._ensure_output_dir(source, opts.get('outputDir', ''), 'excel')

            header, data_rows, active_sheet = self._load_sheet(source, sheet_name, opts)
            schema = self._normalize_schema(schema_text, delimiter, header)
            records = self._rows_to_dicts(schema, data_rows, active_sheet)

            merge_files = opts.get('mergeFiles') or []
            if merge_files:
                merged = self._load_merge_tables(schema, merge_files, opts)
                records.extend(merged)

            records = self._clean_records(records, opts)
            if sort_by and sort_by not in schema:
                raise ValueError(f'排序字段不存在：{sort_by}')
            if group_by and group_by not in schema:
                raise ValueError(f'分组字段不存在：{group_by}')
            if sort_by and sort_by in schema:
                reverse = sort_order == 'desc'
                nonempty = [item for item in records if item.get(sort_by) is not None and item.get(sort_by) != '']
                empty = [item for item in records if item.get(sort_by) is None or item.get(sort_by) == '']
                records = sorted(nonempty, key=lambda item: self._sort_key(item.get(sort_by)), reverse=reverse) + empty

            grouped = self._group_records(records, group_by)
            chart = self._build_chart(grouped) if group_by else {}
            group_files: List[str] = []
            if group_by and export_groups:
                group_dir = output_dir / 'groups'
                group_dir.mkdir(parents=True, exist_ok=True)
                group_files = self._write_groups(schema, grouped, group_dir, source.stem)

            json_path = ''
            if group_by and export_json and chart:
                json_path = self._write_chart_json(chart, output_dir, source.stem)

            combined_path = ''
            if export_combined:
                combined_path = str(output_dir / f'{source.stem}_combined.xlsx')
                combined_path = str(self._write_rows(schema, records, Path(combined_path)))

            summary = {
                'totalRows': len(records),
                'groupBy': group_by,
                'groupCount': len(grouped) if group_by else 0,
                'sortBy': sort_by,
                'sortOrder': sort_order
            }

            return {
                'code': 0,
                'msg': 'Excel 处理完成',
                'summary': summary,
                'schema': schema,
                'groups': [{'key': key, 'count': len(val)} for key, val in grouped.items()] if group_by else [],
                'groupFiles': group_files,
                'jsonPath': json_path,
                'chart': chart,
                'combinedPath': combined_path,
                'outputDir': str(output_dir),
                'sample': self._public_rows(records[:30]),
                'formulaPolicy': opts.get('formulaPolicy') or 'preserve'
            }
        except Exception as exc:
            return {'code': -1, 'msg': f'处理失败：{exc}'}

    @staticmethod
    def _clean_records(records, options):
        cleaned, seen = [], set()
        keys = options.get('deduplicateColumns') or []
        if records and any(key not in records[0] or key.startswith('_') for key in keys):
            raise ValueError('去重字段不存在，请重新选择字段')
        for record in records:
            row = {**record, '_cells': dict(record.get('_cells', {}))}
            if options.get('trimText'):
                for key, value in row.items():
                    if not key.startswith('_') and isinstance(value, str) and row['_cells'].get(key, {}).get('type') != 'f':
                        row[key] = value.strip()
            if keys:
                signature = tuple((type(row.get(key)).__name__, str(row.get(key))) for key in keys)
                if signature in seen:
                    continue
                seen.add(signature)
            cleaned.append(row)
        return cleaned

    def excel_process_preview(self, options=None):
        """Bounded before/after sample; never runs the export or scans the full sheet."""
        opts = self._validate_payload(options)
        preview = self.excel_preview({**opts, 'limit': min(100, int(opts.get('limit') or 30))})
        if preview.get('code') != 0:
            return preview
        before = preview['sample']
        after = self._clean_records(before, opts)
        sort_by = opts.get('sortBy')
        if sort_by:
            nonempty = [row for row in after if row.get(sort_by) is not None and row.get(sort_by) != '']
            empty = [row for row in after if row.get(sort_by) is None or row.get(sort_by) == '']
            after = sorted(nonempty, key=lambda row: self._sort_key(row.get(sort_by)), reverse=opts.get('sortOrder') == 'desc') + empty
        return {**preview, 'before': before, 'after': self._public_rows(after), 'sampleOnly': True,
                'msg': '仅比较当前样本，完整排序和去重将在执行时进行'}

    def excel_merge_tables(self, options: Dict = None):
        '''将多个分表合并为主表'''
        try:
            opts = self._validate_payload(options)
            tables = opts.get('tables') or []
            if not tables:
                raise ValueError('请至少选择一个分表文件')
            first = tables[0]
            schema_text = opts.get('schemaText', '')
            delimiter = opts.get('delimiter') or '|'

            first_path = self._ensure_excel_file(first.get('path', ''))
            header, _, _ = self._load_sheet(first_path, first.get('sheet'), opts, limit=1)
            schema = self._normalize_schema(schema_text, delimiter, header)

            merged: List[Dict[str, Any]] = []
            for table in tables:
                table_path = self._ensure_excel_file(table.get('path', ''))
                sheet_name = table.get('sheet')
                source_header, rows, sheet = self._load_sheet(table_path, sheet_name, {**opts, **table})
                mapping = table.get('fieldMapping') or opts.get('fieldMapping') or {}
                source_records = self._rows_to_dicts(source_header, rows, sheet)
                for row in source_records:
                    mapped = {'_source': sheet, '_cells': {}}
                    for column in schema:
                        source_column = mapping.get(column, column)
                        if source_column not in source_header:
                            raise ValueError(f'{table_path.name} 缺少字段 {source_column}，请设置字段映射')
                        mapped[column] = row.get(source_column)
                        mapped['_cells'][column] = row['_cells'].get(source_column, {})
                    merged.append(mapped)

            output_dir = self._ensure_output_dir(first_path, opts.get('outputDir', ''), 'merged')
            filename = opts.get('outputName') or f'merged_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            dest = output_dir / (filename if filename.lower().endswith('.xlsx') else f'{filename}.xlsx')
            dest = self._write_rows(schema, merged, dest)

            return {
                'code': 0,
                'msg': f'已合并 {len(tables)} 个分表',
                'output': str(dest),
                'rows': len(merged),
                'schema': schema
            }
        except Exception as exc:
            return {'code': -1, 'msg': f'合并失败：{exc}'}

    def excel_column_profile(self, options: Dict = None):
        '''列统计分析'''
        try:
            opts = self._validate_payload(options)
            source = self._ensure_excel_file(opts.get('filePath', ''))
            sheet_name = opts.get('sheetName')
            delimiter = opts.get('delimiter') or '|'
            schema_text = opts.get('schemaText', '')
            header, data_rows, active_sheet = self._load_sheet(source, sheet_name, opts)
            schema = self._normalize_schema(schema_text, delimiter, header)
            target_columns = opts.get('columns') or schema
            columns = [col for col in target_columns if col in schema]
            if not columns:
                columns = schema
            records = self._rows_to_dicts(schema, data_rows, active_sheet)
            total_rows = len(records)
            profiles = []
            for column in columns:
                values = [self._normalize_cell(row.get(column)) for row in records]
                blanks = sum(1 for value in values if not value)
                non_blank = [value for value in values if value]
                unique_count = len(set(non_blank))
                numeric_values = []
                for value in non_blank:
                    try:
                        numeric_values.append(float(value.replace(',', '')))
                    except Exception:
                        continue
                is_numeric = len(numeric_values) >= max(1, int(0.6 * len(non_blank))) if non_blank else False
                counter = Counter(non_blank)
                top_values = [
                    {'value': key, 'count': count, 'ratio': round(count / total_rows, 4)}
                    for key, count in counter.most_common(5)
                ]
                numeric_summary = {}
                if numeric_values:
                    numeric_summary = {
                        'min': min(numeric_values),
                        'max': max(numeric_values),
                        'avg': round(sum(numeric_values) / len(numeric_values), 4)
                    }
                profiles.append({
                    'field': column,
                    'type': 'number' if is_numeric else 'text',
                    'unique': unique_count,
                    'blanks': blanks,
                    'blankRatio': round(blanks / total_rows, 4) if total_rows else 0,
                    'topValues': top_values,
                    'numeric': numeric_summary,
                    'samples': non_blank[:5]
                })
            return {
                'code': 0,
                'msg': '列分析完成',
                'summary': {
                    'sheet': active_sheet,
                    'totalRows': total_rows,
                    'columns': len(columns)
                },
                'profiles': profiles
            }
        except Exception as exc:
            return {'code': -1, 'msg': f'分析失败：{exc}'}

    def excel_quality_report(self, options: Dict = None):
        '''生成可交付的 Excel 数据质量报告。'''
        try:
            opts = self._validate_payload(options)
            source = self._ensure_excel_file(opts.get('filePath', ''))
            profile_result = self.excel_column_profile(opts)
            if profile_result.get('code') != 0:
                raise ValueError(profile_result.get('msg') or '列分析失败')
            summary = profile_result['summary']
            profiles = profile_result['profiles']
            threshold = max(0, min(float(opts.get('blankRatioThreshold') or 0.1), 1))
            output_dir = self._ensure_output_dir(source, opts.get('outputDir', ''), 'quality')
            filename = str(opts.get('outputName') or f'{source.stem}_quality_report.xlsx')
            if not filename.lower().endswith('.xlsx'):
                filename += '.xlsx'
            dest = output_dir / Path(filename).name

            workbook = Workbook()
            overview = workbook.active
            overview.title = '质量概览'
            overview.append(['指标', '值'])
            overview.append(['文件', source.name])
            overview.append(['工作表', summary['sheet']])
            overview.append(['数据行数', summary['totalRows']])
            overview.append(['字段数', summary['columns']])
            overview.append(['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])

            details = workbook.create_sheet('字段画像')
            details.append(['字段', '推断类型', '唯一值', '空值', '空值率', '最小值', '最大值', '平均值', '高频值'])
            issues = workbook.create_sheet('质量问题')
            issues.append(['级别', '字段', '问题', '建议'])
            issue_count = 0
            for profile in profiles:
                numeric = profile.get('numeric') or {}
                top_values = '；'.join(f'{item["value"]} ({item["count"]})' for item in profile.get('topValues') or [])
                details.append([
                    profile['field'],
                    profile['type'],
                    profile['unique'],
                    profile['blanks'],
                    profile['blankRatio'],
                    numeric.get('min', ''),
                    numeric.get('max', ''),
                    numeric.get('avg', ''),
                    top_values,
                ])
                if profile['blankRatio'] >= threshold and profile['blanks']:
                    issues.append(['警告', profile['field'], f'空值率 {profile["blankRatio"]:.1%}', '检查缺失数据来源，确定填充或删除策略'])
                    issue_count += 1
                non_blank = max(0, summary['totalRows'] - profile['blanks'])
                duplicate_count = max(0, non_blank - profile['unique'])
                if duplicate_count and profile['unique'] == 1:
                    issues.append(['提示', profile['field'], '除空值外仅有一个值', '确认该字段是否仍有分析价值'])
                    issue_count += 1
            for sheet in (overview, details, issues):
                sheet.freeze_panes = 'A2'
                for column in sheet.columns:
                    width = min(60, max(10, max((len(str(cell.value or '')) for cell in column), default=8) + 2))
                    sheet.column_dimensions[column[0].column_letter].width = width
            dest = write_output(dest, lambda target: workbook.save(target))
            workbook.close()
            return {
                'code': 0,
                'msg': f'数据质量报告已生成，发现 {issue_count} 项提示',
                'output': str(dest),
                'outputDir': str(output_dir),
                'issueCount': issue_count,
                'summary': summary,
                'profiles': profiles,
            }
        except Exception as exc:
            return {'code': -1, 'msg': f'生成质量报告失败：{exc}'}

    def excel_split_by_column(self, options: Dict = None):
        '''按列拆分为多个工作簿'''
        try:
            opts = self._validate_payload(options)
            source = self._ensure_excel_file(opts.get('filePath', ''))
            column = opts.get('column') or opts.get('groupBy')
            if not column:
                raise ValueError('请选择需要拆分的列')
            sheet_name = opts.get('sheetName')
            delimiter = opts.get('delimiter') or '|'
            schema_text = opts.get('schemaText', '')
            header, data_rows, active_sheet = self._load_sheet(source, sheet_name, opts)
            schema = self._normalize_schema(schema_text, delimiter, header)
            if column not in schema:
                raise ValueError(f'列不存在：{column}')
            records = self._rows_to_dicts(schema, data_rows, active_sheet)
            limit = int(opts.get('limit') or 0)
            min_rows = int(opts.get('minRows') or 1)
            empty_label = opts.get('emptyLabel') or '未分类'
            output_dir = self._ensure_output_dir(source, opts.get('outputDir', ''), 'split')
            groups = self._group_records(records, column, empty_label)
            sorted_groups = sorted(groups.items(), key=lambda item: len(item[1]), reverse=True)
            if limit > 0:
                sorted_groups = sorted_groups[:limit]
            exports = []
            for label, rows in sorted_groups:
                if len(rows) < min_rows:
                    continue
                safe_label = ''.join(ch if ch.isalnum() else '_' for ch in label) or 'group'
                dest = output_dir / f'{self._safe_filename_part(column, "column")}_{safe_label}.xlsx'
                dest = self._write_rows(schema, rows, dest)
                exports.append({'label': label, 'rows': len(rows), 'file': str(dest)})
            if not exports:
                raise ValueError('没有满足条件的分组可导出')
            return {
                'code': 0,
                'msg': f'已导出 {len(exports)} 个分组',
                'groups': [{'label': item['label'], 'rows': item['rows']} for item in exports],
                'files': [item['file'] for item in exports],
                'outputDir': str(output_dir)
            }
        except Exception as exc:
            return {'code': -1, 'msg': f'拆分失败：{exc}'}

    # -------- internal helpers --------

    def _list_sheets(self, file_path: Path) -> List[str]:
        wb = load_workbook(file_path, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names

    def _group_records(self, rows: List[Dict[str, Any]], group_by: str, empty_label='') -> Dict[str, List[Dict[str, Any]]]:
        if not group_by:
            return {}
        typed = defaultdict(list)
        for row in rows:
            value = row.get(group_by)
            label = self._normalize_cell(value) if value is not None and value != '' else empty_label
            kind = ('空值' if value is None else '布尔' if isinstance(value, bool) else
                    '数字' if isinstance(value, (int, float)) else '日期' if isinstance(value, (date, datetime)) else '文本')
            typed[(kind, label)].append(row)
        labels = defaultdict(int)
        for _, label in typed:
            labels[label] += 1
        grouped = {}
        for (kind, label), group in typed.items():
            display = f'{label}（{kind}）' if labels[label] > 1 else label
            unique, index = display, 2
            while unique in grouped:
                unique, index = f'{display}_{index}', index + 1
            grouped[unique] = group
        return grouped

    def _load_merge_tables(self, schema: Sequence[str], tables: List[Dict[str, Any]], options=None) -> List[Dict[str, Any]]:
        merged: List[Dict[str, Any]] = []
        for table in tables:
            table_path = self._ensure_excel_file(table.get('path', ''))
            sheet_name = table.get('sheet')
            header, rows, sheet = self._load_sheet(table_path, sheet_name, {**(options or {}), **table})
            mapping = table.get('fieldMapping') or {}
            for row in self._rows_to_dicts(header, rows, sheet):
                mapped = {'_source': sheet, '_cells': {}}
                for column in schema:
                    source_column = mapping.get(column, column)
                    if source_column not in header:
                        raise ValueError(f'{table_path.name} 缺少字段 {source_column}，请配置字段映射')
                    mapped[column] = row.get(source_column)
                    mapped['_cells'][column] = row['_cells'].get(source_column, {})
                merged.append(mapped)
        return merged
