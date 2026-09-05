#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Offline OCR for images and scanned PDFs."""
from __future__ import annotations

import csv
import json
import statistics
import threading
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

from api.core.context import checkpoint, iter_progress
from api.core.outputs import atomic_output, write_output
from api.utils.error_handler import api_success, safe_execute
from pyapp.config.config import Config


class OcrMixin:
    """RapidOCR-backed local text recognition with optional searchable PDF output."""

    _ocr_engine = None
    _ocr_lock = threading.Lock()

    @classmethod
    def _get_ocr_engine(cls):
        if cls._ocr_engine is not None:
            return cls._ocr_engine
        with cls._ocr_lock:
            if cls._ocr_engine is None:
                try:
                    from rapidocr import RapidOCR
                except ImportError as exc:
                    raise RuntimeError(f'离线 OCR 组件未安装，请重新安装 {Config.appVersion} 完整版') from exc
                cls._ocr_engine = RapidOCR()
        return cls._ocr_engine

    @staticmethod
    def _validate_options(options: Dict | None) -> Dict:
        if not isinstance(options, dict):
            raise ValueError('参数格式错误')
        return options

    @staticmethod
    def _source_path(options: Dict, allowed: Sequence[str]) -> Path:
        raw = options.get('filePath') or options.get('file')
        if isinstance(raw, dict):
            raw = raw.get('path')
        if not raw:
            raise ValueError('请选择源文件')
        source = Path(str(raw)).expanduser().resolve()
        if not source.is_file():
            raise FileNotFoundError(f'文件不存在：{source}')
        if source.suffix.lower() not in allowed:
            raise ValueError(f'不支持的文件格式：{source.suffix or "未知"}')
        return source

    @staticmethod
    def _unique_output(source: Path, options: Dict, marker: str, extension: str) -> Path:
        raw_dir = options.get('outputDir')
        output_dir = Path(str(raw_dir)).expanduser() if raw_dir else source.parent
        output_dir.mkdir(parents=True, exist_ok=True)
        custom_name = str(options.get('outputName') or '').strip()
        if custom_name:
            filename = Path(custom_name).name
            if not filename.lower().endswith(extension.lower()):
                filename += extension
        else:
            filename = f'{source.stem}_{marker}{extension}'
        target = output_dir / filename
        index = 2
        while target.exists() or target.is_symlink():
            target = output_dir / f'{Path(filename).stem}_{index}{extension}'
            index += 1
        return target

    @staticmethod
    def _parse_result(result) -> List[Dict]:
        if result is None:
            return []
        boxes = getattr(result, 'boxes', None)
        texts = getattr(result, 'txts', None)
        scores = getattr(result, 'scores', None)
        if texts is None and isinstance(result, (tuple, list)) and result:
            rows = result[0] if len(result) == 2 else result
            parsed = []
            for row in rows or []:
                if not isinstance(row, (tuple, list)) or len(row) < 2:
                    continue
                box = row[0]
                text_score = row[1]
                text = text_score[0] if isinstance(text_score, (tuple, list)) else text_score
                score = text_score[1] if isinstance(text_score, (tuple, list)) and len(text_score) > 1 else 0
                parsed.append({'box': box, 'text': str(text), 'score': float(score or 0)})
            return parsed
        texts = list(texts) if texts is not None else []
        boxes = list(boxes) if boxes is not None else [None] * len(texts)
        scores = list(scores) if scores is not None else [0] * len(texts)
        return [
            {'box': (boxes[index].tolist() if hasattr(boxes[index], 'tolist') else boxes[index]) if index < len(boxes) else None, 'text': str(text), 'score': float(scores[index] or 0)}
            for index, text in enumerate(texts)
            if str(text).strip()
        ]

    @classmethod
    def _recognize(cls, image, correct_line_direction=True) -> List[Dict]:
        import numpy as np

        engine = cls._get_ocr_engine()
        with cls._ocr_lock:
            result = engine(np.asarray(image.convert('RGB')), use_cls=correct_line_direction)
        return cls._parse_result(result)

    @classmethod
    def _recognize_oriented(cls, image, options):
        from PIL import ImageOps
        oriented = ImageOps.exif_transpose(image)
        explicit = int(options.get('rotation') or 0) % 360
        angles = [explicit] if not options.get('autoRotate', False) else [0, 90, 180, 270]
        candidates = []
        for angle in angles:
            checkpoint()
            working = oriented.rotate(angle, expand=True) if angle else oriented
            # Automatic line classification hides an upside-down page by rotating
            # individual crops. Disable it while comparing page orientations.
            lines = cls._recognize(working, correct_line_direction=not options.get('autoRotate', False))
            score = 0
            for line in lines:
                box = line.get('box')
                horizontal = 1
                if box and len(box) >= 4:
                    width = max(point[0] for point in box) - min(point[0] for point in box)
                    height = max(point[1] for point in box) - min(point[1] for point in box)
                    horizontal = min(1, width / max(1, height))
                score += len(line['text']) * max(0, line['score']) ** 2 * horizontal
            candidates.append((score, angle, working.size, lines))
        _, angle, size, lines = max(candidates, key=lambda candidate: candidate[0])
        threshold = max(0, min(1, float(options.get('confidenceThreshold', 0.8))))
        for line in lines:
            line['lowConfidence'] = line['score'] < threshold
        return lines, angle, size

    @staticmethod
    def _summary(lines: Iterable[Dict]) -> Dict:
        rows = list(lines)
        text = '\n'.join(row['text'] for row in rows)
        average = sum(row['score'] for row in rows) / len(rows) if rows else 0
        return {
            'text': text,
            'preview': text[:4000],
            'lineCount': len(rows),
            'averageConfidence': round(average, 4),
            'lines': rows,
            'lowConfidenceCount': sum(row.get('lowConfidence', row['score'] < 0.8) for row in rows),
        }

    @staticmethod
    def _parse_pages(page_spec: str, total: int) -> List[int]:
        if not page_spec or not page_spec.strip():
            return list(range(total))
        pages = []
        for part in page_spec.replace('，', ',').split(','):
            token = part.strip()
            if not token:
                continue
            if '-' in token:
                start_text, end_text = token.split('-', 1)
                start, end = int(start_text), int(end_text)
                if start > end:
                    start, end = end, start
                pages.extend(range(start - 1, end))
            else:
                pages.append(int(token) - 1)
        unique = []
        for page in pages:
            if page < 0 or page >= total:
                raise ValueError(f'页码超出范围：{page + 1}（共 {total} 页）')
            if page not in unique:
                unique.append(page)
        if not unique:
            raise ValueError('未选择有效页码')
        return unique

    @safe_execute
    def ocr_image(self, options=None):
        from PIL import Image

        opts = self._validate_options(options)
        source = self._source_path(opts, ('.png', '.jpg', '.jpeg', '.webp', '.bmp', '.tif', '.tiff'))
        with Image.open(source) as image:
            lines, rotation, _ = self._recognize_oriented(image, opts)
        summary = self._summary(lines)
        output = ''
        if bool(opts.get('saveFile', True)):
            target = self._unique_output(source, opts, 'ocr', '.txt')
            target = write_output(target, lambda path: path.write_text(summary['text'], encoding='utf-8'))
            output = str(target)
        return api_success('OCR 识别完成', output=output, rotation=rotation, **summary)

    @staticmethod
    def _insert_searchable_text(page, lines: List[Dict], pixel_width: int, pixel_height: int) -> None:
        import fitz

        scale_x = page.rect.width / max(1, pixel_width)
        scale_y = page.rect.height / max(1, pixel_height)
        for line in lines:
            box = line.get('box')
            text = line.get('text', '').strip()
            if box is None or not text:
                continue
            try:
                xs = [float(point[0]) for point in box]
                ys = [float(point[1]) for point in box]
                rect = fitz.Rect(min(xs) * scale_x, min(ys) * scale_y, max(xs) * scale_x, max(ys) * scale_y)
                if rect.is_empty or rect.is_infinite:
                    continue
                font_size = max(4.0, min(24.0, rect.height * 0.8))
                page.insert_textbox(rect, text, fontsize=font_size, fontname='china-s', render_mode=3, overlay=True)
            except Exception:
                # A malformed OCR box should not make the whole document fail.
                continue

    @safe_execute
    def ocr_pdf(self, options=None):
        import fitz
        from PIL import Image

        opts = self._validate_options(options)
        source = self._source_path(opts, ('.pdf',))
        dpi = max(120, min(400, int(opts.get('dpi') or 220)))
        mode = str(opts.get('outputMode') or 'both').lower()
        if mode not in {'text', 'searchable_pdf', 'both'}:
            raise ValueError('输出模式仅支持 text、searchable_pdf 或 both')

        source_doc = fitz.open(source)
        output_doc = None
        try:
            pages = self._parse_pages(str(opts.get('pageSpec') or ''), source_doc.page_count)
            if mode in {'searchable_pdf', 'both'}:
                output_doc = fitz.open()
                output_doc.insert_pdf(source_doc)
            page_results = []
            matrix = fitz.Matrix(dpi / 72, dpi / 72)
            for page_index in iter_progress(pages, '正在识别 PDF 页面'):
                page = source_doc.load_page(page_index)
                pixmap = page.get_pixmap(matrix=matrix, alpha=False)
                image = Image.frombytes('RGB', (pixmap.width, pixmap.height), pixmap.samples)
                lines, rotation, _ = self._recognize_oriented(image, opts)
                page_results.append({'page': page_index + 1, 'rotation': rotation, 'lines': lines})
                if output_doc is not None:
                    mapped_lines = []
                    for line in lines:
                        points = []
                        for x, y in line.get('box') or []:
                            points.append([pixmap.width - y, x] if rotation == 90 else [pixmap.width - x, pixmap.height - y] if rotation == 180 else [y, pixmap.height - x] if rotation == 270 else [x, y])
                        mapped_lines.append({**line, 'box': points})
                    self._insert_searchable_text(output_doc.load_page(page_index), mapped_lines, pixmap.width, pixmap.height)

            text = '\n\n'.join(
                f'--- 第 {item["page"]} 页 ---\n' + '\n'.join(line['text'] for line in item['lines'])
                for item in page_results
            )
            scores = [line['score'] for item in page_results for line in item['lines']]
            outputs = []
            text_output = ''
            pdf_output = ''
            if mode in {'text', 'both'}:
                text_target = self._unique_output(source, opts, 'ocr', '.txt')
                text_target = write_output(text_target, lambda target: target.write_text(text, encoding='utf-8'))
                text_output = str(text_target)
                outputs.append(text_output)
            if output_doc is not None:
                pdf_target = self._unique_output(source, opts, 'searchable', '.pdf')
                pdf_target = write_output(pdf_target, lambda target: output_doc.save(target, garbage=4, deflate=True))
                pdf_output = str(pdf_target)
                outputs.append(pdf_output)
            return api_success(
                'PDF OCR 完成',
                output=pdf_output or text_output,
                outputs=outputs,
                textOutput=text_output,
                pdfOutput=pdf_output,
                preview=text[:4000],
                pageCount=len(pages),
                lineCount=len(scores),
                averageConfidence=round(sum(scores) / len(scores), 4) if scores else 0,
                pages=page_results, lowConfidenceCount=sum(line.get('lowConfidence', False) for item in page_results for line in item['lines']),
            )
        finally:
            if output_doc is not None:
                output_doc.close()
            source_doc.close()

    @staticmethod
    def _table_from_lines(lines: List[Dict], image_width: int, tolerance: float = 0) -> List[List[str]]:
        cells = []
        for line in lines:
            box = line.get('box')
            text = str(line.get('text') or '').strip()
            if not text or not isinstance(box, (list, tuple)) or len(box) < 2:
                continue
            try:
                xs = [float(point[0]) for point in box]
                ys = [float(point[1]) for point in box]
            except (TypeError, ValueError, IndexError):
                continue
            x0, x1 = min(xs), max(xs)
            y0, y1 = min(ys), max(ys)
            cells.append({'text': text, 'x0': x0, 'x1': x1, 'cy': (y0 + y1) / 2, 'height': max(1, y1 - y0)})
        if not cells:
            return []

        median_height = statistics.median(cell['height'] for cell in cells)
        row_tolerance = max(4.0, median_height * 0.65)
        grouped_rows = []
        for cell in sorted(cells, key=lambda item: (item['cy'], item['x0'])):
            row = next((candidate for candidate in grouped_rows if abs(candidate['cy'] - cell['cy']) <= row_tolerance), None)
            if row is None:
                grouped_rows.append({'cy': cell['cy'], 'cells': [cell]})
            else:
                row['cells'].append(cell)
                row['cy'] = sum(item['cy'] for item in row['cells']) / len(row['cells'])

        column_tolerance = float(tolerance or max(18, image_width * 0.035))
        anchors: List[Dict] = []
        for x0 in sorted(cell['x0'] for cell in cells):
            anchor = next((item for item in anchors if abs(item['x'] - x0) <= column_tolerance), None)
            if anchor is None:
                anchors.append({'x': x0, 'values': [x0]})
            else:
                anchor['values'].append(x0)
                anchor['x'] = sum(anchor['values']) / len(anchor['values'])
        anchors.sort(key=lambda item: item['x'])

        table: List[List[str]] = []
        for row in sorted(grouped_rows, key=lambda item: item['cy']):
            values = [''] * len(anchors)
            for cell in sorted(row['cells'], key=lambda item: item['x0']):
                column = min(range(len(anchors)), key=lambda index: abs(anchors[index]['x'] - cell['x0']))
                values[column] = f'{values[column]} {cell["text"]}'.strip()
            while values and not values[-1]:
                values.pop()
            table.append(values)

        used_columns = [index for index in range(len(anchors)) if any(index < len(row) and row[index] for row in table)]
        return [[row[index] if index < len(row) else '' for index in used_columns] for row in table]

    @safe_execute
    def ocr_table(self, options=None):
        '''从图片或 PDF 中识别规则表格，并导出 CSV / XLSX / JSON。'''
        import fitz
        from openpyxl import Workbook
        from PIL import Image

        opts = self._validate_options(options)
        source = self._source_path(opts, ('.png', '.jpg', '.jpeg', '.webp', '.bmp', '.tif', '.tiff', '.pdf'))
        output_format = str(opts.get('outputFormat') or 'xlsx').lower()
        if output_format not in {'csv', 'xlsx', 'json', 'all'}:
            raise ValueError('输出格式仅支持 csv、xlsx、json 或 all')
        column_tolerance = max(0, float(opts.get('columnTolerance') or 0))
        page_tables = opts.get('tables') or []

        if page_tables:
            if not isinstance(page_tables, list) or len(page_tables) > 10000:
                raise ValueError('表格页数据格式无效')
            for item in page_tables:
                if not isinstance(item, dict) or not isinstance(item.get('rows'), list):
                    raise ValueError('每页必须包含 rows 列表')
                for row in item['rows']:
                    if not isinstance(row, list) or any(not isinstance(value, (str, int, float, bool, type(None))) for value in row):
                        raise ValueError('表格单元格必须为文本或数值')
        elif source.suffix.lower() == '.pdf':
            dpi = max(120, min(400, int(opts.get('dpi') or 220)))
            with fitz.open(source) as doc:
                pages = self._parse_pages(str(opts.get('pageSpec') or ''), doc.page_count)
                matrix = fitz.Matrix(dpi / 72, dpi / 72)
                for page_index in iter_progress(pages, '正在识别表格'):
                    page = doc.load_page(page_index)
                    pixmap = page.get_pixmap(matrix=matrix, alpha=False)
                    image = Image.frombytes('RGB', (pixmap.width, pixmap.height), pixmap.samples)
                    lines, rotation, size = self._recognize_oriented(image, opts)
                    table = self._table_from_lines(lines, size[0], column_tolerance)
                    page_tables.append({'page': page_index + 1, 'rows': table, 'rotation': rotation, 'uncertain': [line['text'] for line in lines if line['lowConfidence']]})
        else:
            with Image.open(source) as image:
                lines, rotation, size = self._recognize_oriented(image, opts)
                page_tables.append({'page': 1, 'rows': self._table_from_lines(lines, size[0], column_tolerance), 'rotation': rotation, 'uncertain': [line['text'] for line in lines if line['lowConfidence']]})

        if not any(item['rows'] for item in page_tables):
            raise ValueError('未识别到可导出的表格内容')
        if opts.get('saveFile') is False:
            return api_success('表格已识别，请核对低置信度内容后导出', tables=page_tables, preview=page_tables,
                               outputs=[], rowCount=sum(len(item['rows']) for item in page_tables),
                               columnCount=max((len(row) for item in page_tables for row in item['rows']), default=0))
        output_dir_raw = opts.get('outputDir')
        output_dir = Path(str(output_dir_raw)).expanduser() if output_dir_raw else source.parent
        output_dir.mkdir(parents=True, exist_ok=True)
        base_name = Path(str(opts.get('outputName') or f'{source.stem}_table')).stem
        outputs = []

        if output_format in {'csv', 'all'}:
            for item in page_tables:
                suffix = f'_p{item["page"]}' if len(page_tables) > 1 else ''
                target = output_dir / f'{base_name}{suffix}.csv'
                index = 2
                while target.exists() or target.is_symlink():
                    target = output_dir / f'{base_name}{suffix}_{index}.csv'
                    index += 1
                with atomic_output(target) as (temporary, final):
                    with temporary.open('w', encoding='utf-8-sig', newline='') as handler:
                        writer = csv.writer(handler)
                        writer.writerows(item['rows'])
                target = final
                outputs.append(str(target))

        if output_format in {'xlsx', 'all'}:
            target = self._unique_output(source, {**opts, 'outputName': base_name}, 'table', '.xlsx')
            workbook = Workbook()
            workbook.remove(workbook.active)
            for item in page_tables:
                sheet = workbook.create_sheet(title=f'第{item["page"]}页')
                for row in item['rows']:
                    sheet.append(row)
                    for cell in sheet[sheet.max_row]:
                        if isinstance(cell.value, str):
                            cell.data_type = 's'
                for column in sheet.columns:
                    width = min(60, max(10, max((len(str(cell.value or '')) for cell in column), default=8) + 2))
                    sheet.column_dimensions[column[0].column_letter].width = width
            target = write_output(target, lambda path: workbook.save(path))
            workbook.close()
            outputs.append(str(target))

        if output_format in {'json', 'all'}:
            target = self._unique_output(source, {**opts, 'outputName': base_name}, 'table', '.json')
            target = write_output(target, lambda path: path.write_text(json.dumps(page_tables, ensure_ascii=False, indent=2), encoding='utf-8'))
            outputs.append(str(target))

        row_count = sum(len(item['rows']) for item in page_tables)
        column_count = max((len(row) for item in page_tables for row in item['rows']), default=0)
        return api_success(
            f'表格识别完成，共 {row_count} 行、{column_count} 列',
            output=outputs[0] if outputs else '',
            outputs=outputs,
            outputDir=str(output_dir),
            rowCount=row_count,
            columnCount=column_count,
            preview=page_tables[:3],
        )
